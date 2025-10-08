from fastapi import APIRouter, UploadFile, File, HTTPException, Depends
from sqlalchemy.orm import Session
from openpyxl import load_workbook
from typing import List
from database import SessionLocal
from app.schema.products import ProductsResponse, ProductsCreate
from app.controllers.products import upload_products
import os

router = APIRouter()

# ---------- DB Dependency ----------
def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

# ---------- Utility Cleaners ----------
def clean_value(val):
    """Convert empty/invalid to None, preserve 'N/A'."""
    if val is None:
        return None
    val_str = str(val).replace("\xa0", " ").strip()  # remove non-breaking spaces
    if val_str == "" or val_str.lower() in ["none", "null", "-", "na"]:
        return None
    if val_str.lower() == "n/a":
        return "N/A"
    return val_str


def clean_numeric(val, default=0.0):
    """Convert to float safely."""
    if val is None:
        return default
    val_str = str(val).replace("\xa0", " ").strip()
    if val_str == "" or val_str.lower() in ["n/a", "na", "none", "-"]:
        return None
    val_str = (
        val_str.replace("₹", "")
        .replace("$", "")
        .replace("€", "")
        .replace("£", "")
        .strip()
    )
    try:
        return float(val_str)
    except ValueError:
        return default


def clean_int(val, default=0):
    """Convert to int safely."""
    if val is None:
        return default
    val_str = str(val).replace("\xa0", " ").strip()
    if val_str == "" or val_str.lower() in ["n/a", "na", "none", "-"]:
        return default
    try:
        return int(float(val_str))
    except ValueError:
        return default


def clean_bool(val):
    """Convert Excel bool-like values to True/False."""
    if val is None:
        return True  # default True
    val_str = str(val).replace("\xa0", " ").strip().lower()
    if val_str in ["true", "yes", "1", "y"]:
        return True
    if val_str in ["false", "no", "0", "n"]:
        return False
    return True  # fallback default


# ---------- Main Import Route ----------
@router.post("/import-products/", response_model=List[ProductsResponse])
async def import_products(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """
    Imports products from Excel into DB using header-based column mapping.
    Expected headers:
    itemcode, itemname, description, brand, watt, color, cct, beamangle, cri, lumens,
    price, quantity, unit, rackcode, size, cutoutdia, category, subcategory,
    in_display, model, reorderqty
    """

    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Invalid file format. Please upload an Excel file.")

    try:
        # Save temporary Excel file
        contents = await file.read()
        temp_file_path = "temp_import.xlsx"
        with open(temp_file_path, "wb") as f:
            f.write(contents)

        workbook = load_workbook(filename=temp_file_path, data_only=True)
        sheet = workbook.active

        # Get headers (lowercase, stripped)
        headers = [str(cell.value).strip().lower() if cell.value else "" for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        print("🧾 Excel Headers:", headers)

        products_list = []

        for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if all(cell is None for cell in row):
                continue

            row_data = dict(zip(headers, row))

            product_data_dict = {
                "itemcode": clean_value(row_data.get("itemcode")),
                "itemname": clean_value(row_data.get("itemname")),
                "description": clean_value(row_data.get("description")),
                "brand": clean_value(row_data.get("brand")),
                "watt": clean_value(row_data.get("watt")),
                "color": clean_value(row_data.get("color")),
                "cct": clean_value(row_data.get("cct")),
                "beamangle": clean_value(row_data.get("beamangle")),
                "cri": clean_value(row_data.get("cri")),
                "lumens": clean_value(row_data.get("lumens")),
                "price": clean_numeric(row_data.get("price"), 0.0),
                "quantity": clean_int(row_data.get("quantity"), 0),
                "unit": clean_value(row_data.get("unit")),
                "rackcode": clean_value(row_data.get("rackcode")),
                "size": clean_value(row_data.get("size")),
                "cutoutdia": clean_value(row_data.get("cutoutdia")),
                "category": clean_value(row_data.get("category")),
                "subcategory": clean_value(row_data.get("subcategory")),
                "in_display": clean_bool(row_data.get("in_display")),
                "model": clean_value(row_data.get("model")),
                "reorderqty": clean_int(row_data.get("reorderqty"), 10),
            }

            # Skip rows missing required fields
            if not product_data_dict["itemcode"] or not product_data_dict["itemname"]:
                print(f"⚠️ Skipped row {idx}: Missing itemcode or itemname")
                continue

            # Debug for first 3 rows
            if idx <= 4:
                print(f"Row {idx} processed:", product_data_dict)

            products_list.append(ProductsCreate(**product_data_dict))

        os.remove(temp_file_path)

        if not products_list:
            raise HTTPException(status_code=400, detail="No valid product data found in the file.")

        created_products = upload_products(products_list, db)
        return created_products

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error processing file: {str(e)}")
